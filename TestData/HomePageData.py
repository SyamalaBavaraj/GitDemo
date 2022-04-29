import openpyxl


class HomePageData:

    test_HomePage_data = [{'firstname': ' Syamala', 'email': 'test@gmail.com', 'password': 'syam@123', 'gender': 'Female'}, {'firstname': ' Yash', 'email': 'test2@yahoo.com', 'password': 'yash@123', 'gender': 'Male'}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\Administrator\\Desktop\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["firstname"] = "Syamala"
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[Dict]